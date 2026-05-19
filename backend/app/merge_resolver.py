from openpyxl.worksheet.worksheet import Worksheet

def build_merge_map(ws: Worksheet) -> dict:
    """
    Returns {(row, col): value} for every cell inside a merged range,
    pointing back to the top-left root cell's value.
    """
    merge_map = {}
    for merge_range in ws.merged_cells.ranges:
        root = ws.cell(merge_range.min_row, merge_range.min_col)
        value = root.value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merge_map[(row, col)] = value
    return merge_map


def get_cell_value(ws: Worksheet, row: int, col: int, merge_map: dict):
    direct = ws.cell(row, col).value
    if direct is not None:
        return direct
    return merge_map.get((row, col))
