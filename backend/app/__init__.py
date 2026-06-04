import asyncio
from openpyxl import load_workbook

from openpyxl.worksheet.worksheet import Worksheet

from .merge_resolver import build_merge_map
from .header_parser import parse_time_headers
from .faculty_mapper import build_faculty_map, resolve_teacher
from .section_regular import parse_regular_section
from .section_lab import parse_lab_section
from .section_online import parse_online_section




def _detect_boundaries(ws: Worksheet):
    """
    Scans column 1 to find section boundaries.
    Returns: (reg_start, reg_end, lab_start, lab_end,
              online_label, online_header, online_start, online_end)
    """
    max_row = ws.max_row
    reg_start = 5
    lab_start = -1
    online_label = -1
    online_header = -1
    online_start = -1

    for r in range(1, max_row + 1):
        val = ws.cell(r, 1).value
        if isinstance(val, str):
            val_upper = val.strip().upper()
            if val_upper == "LAB":
                lab_start = r
            elif "ONLINE" in val_upper:
                online_label = r
                
    reg_end = max_row
    lab_end = max_row
    online_end = max_row

    if online_label > 0:
        online_header = online_label + 2
        online_start = online_label + 3
        
        if lab_start > 0:
            lab_end = online_label - 1
            reg_end = lab_start - 1
        else:
            reg_end = online_label - 1
    elif lab_start > 0:
        reg_end = lab_start - 1
        
    return (reg_start, reg_end, lab_start, lab_end,
            online_label, online_header, online_start, online_end)


def _parse_workbook(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    faculty_map = build_faculty_map(wb)
    
    title = None
    season = None
    odd_week_dates = []
    even_week_dates = []

    # Extract title and dates from the first non-faculty sheet
    for sheet_name in wb.sheetnames:
        if "faculty" in sheet_name.strip().lower():
            continue
        ws = wb[sheet_name]
        
        # 1. Title
        val = ws.cell(1, 1).value
        if val and isinstance(val, str):
            title = val.strip().replace('\n', ' ')
            import re
            match = re.search(r'\b(summer|spring|fall|autumn|winter)\b\W*(\d{4})\b', title, re.IGNORECASE)
            if match:
                season = f"{match.group(1).capitalize()} - {match.group(2)}"
            
        # 2. Dates table
        date_col = -1
        date_start_row = -1
        for r in range(1, ws.max_row + 1):
            for c in range(1, 15):
                cell_val = str(ws.cell(r, c).value or "").strip().lower()
                if cell_val == "date":
                    date_col = c
                    date_start_row = r + 1
                    break
            if date_col != -1:
                break
                
        if date_start_row != -1:
            all_dates = []
            for r in range(date_start_row, ws.max_row + 1):
                d_val = ws.cell(r, date_col).value
                if not d_val:
                    break
                all_dates.append(str(d_val).strip())
                
            for i, d in enumerate(all_dates):
                if i % 2 == 0:
                    odd_week_dates.append(d)
                else:
                    even_week_dates.append(d)
                    
        break  # Only extract metadata from the primary sheet

    raw_entries = []
    for sheet_name in wb.sheetnames:
        if "faculty" in sheet_name.strip().lower():
            continue
            
        ws = wb[sheet_name]
        merge_map = build_merge_map(ws)
        time_slots = parse_time_headers(ws)
        
        # Determine base day
        sheet_name_clean = sheet_name.strip().lower()
        day_mapping = {
            "sunday": "Sunday",
            "monday": "Monday",
            "tuesday": "Tuesday",
            "wednesday": "Wednesday",
            "thursday": "Thursday",
            "friday": "Friday",
            "saturday": "Saturday"
        }
        day = day_mapping.get(sheet_name_clean, sheet_name.strip())
        
        # Check if sheet header contains "ECSE"
        has_ecse = False
        for r in range(1, 5):
            for c in range(1, 11):
                val = ws.cell(r, c).value
                if val and isinstance(val, str) and "ECSE" in val.upper():
                    has_ecse = True
                    break
            if has_ecse:
                break
                
        # Check if explicitly Wednesday
        is_wednesday = (day == "Wednesday")
        if not is_wednesday:
            for r in range(1, 5):
                for c in range(1, 11):
                    val = ws.cell(r, c).value
                    if val and isinstance(val, str) and "WEDNESDAY" in val.upper():
                        is_wednesday = True
                        break
                if is_wednesday:
                    break
                    
        if has_ecse and not is_wednesday:
            day = "Friday"
            
        day_entries = []
        bounds = _detect_boundaries(ws)
        r_start, r_end, l_start, l_end, o_label, o_header, o_start, o_end = bounds
        
        if r_start <= r_end:
            day_entries.extend(parse_regular_section(ws, time_slots, merge_map, r_start, r_end))
        if l_start > 0 and l_start <= l_end:
            day_entries.extend(parse_lab_section(ws, time_slots, merge_map, l_start, l_end))
        if o_start > 0 and o_start <= o_end:
            day_entries.extend(parse_online_section(ws, merge_map, o_label, o_header, o_start, o_end))
            
        for e in day_entries:
            # Online entries may override the sheet day (e.g. Wednesday instead of Friday)
            if "_online_day" in e:
                e["day"] = e.pop("_online_day")
            else:
                e["day"] = day
        raw_entries.extend(day_entries)

    # Process and map fields for Pydantic models
    entries = []
    for e in raw_entries:
        # Resolve teacher acronym → full name
        teacher = resolve_teacher(e["teacher_acro"], faculty_map)
        
        # Split time_slot into start_time & end_time
        parts = [p.strip() for p in e["time_slot"].replace("–", "-").replace("—", "-").split("-")]
        start_time = parts[0]
        end_time = parts[1] if len(parts) > 1 else ""
        
        # Map section_type to type & odd_even
        sec_type = e["section_type"]
        type_val = "lab" if sec_type.startswith("lab") else "theory"
        
        if "_odd" in sec_type:
            odd_even_val = "odd"
        elif "_even" in sec_type:
            odd_even_val = "even"
        else:
            odd_even_val = None

        # Forcefully remove odd/even tag for CSE 1258 at the source in the parser
        if e["course_code"].strip().upper().replace(" ", "") == "CSE1258":
            type_val = "theory"
            odd_even_val = None
            e["week_note"] = ""

        entries.append({
            **e,
            "teacher": teacher,
            "start_time": start_time,
            "end_time": end_time,
            "course": e["course_code"],
            "type": type_val,
            "odd_even": odd_even_val
        })

    # Index by group
    index: dict[str, list] = {}
    for e in entries:
        g = e["group"]
        index.setdefault(g, []).append(e)

    groups = sorted(index.keys())

    return {
        "groups":  groups,
        "index":   index,
        "total":   len(entries),
        "title":   title,
        "season":  season,
        "odd_week_dates": odd_week_dates,
        "even_week_dates": even_week_dates,
    }


async def parse_excel(path: str) -> dict:
    """Non-blocking wrapper — runs CPU-bound parsing on a worker thread."""
    return await asyncio.to_thread(_parse_workbook, path)
