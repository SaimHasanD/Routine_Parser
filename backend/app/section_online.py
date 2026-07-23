from openpyxl.worksheet.worksheet import Worksheet
from .merge_resolver import get_cell_value, build_merge_map
from .header_parser import parse_online_time_headers
from .section_utils import extract_row_entries
import re


ONLINE_ROOM        = "Online"
ODD_MARKER         = "odd"
EVEN_MARKER        = "even"


def _detect_online_days(ws: Worksheet, label_row: int) -> list[str]:
    """Parse the online section label row (e.g. 'Online Classes (Saturday/ Wednesday)')
    to determine what days online classes fall on (in order of appearance)."""
    if label_row <= 0: return []
    for c in range(1, 15):
        val = ws.cell(label_row, c).value
        if val and isinstance(val, str) and "online" in val.lower():
            text = val.lower()
            day_map = {
                "saturday": "Saturday",
                "sunday": "Sunday",
                "monday": "Monday",
                "tuesday": "Tuesday",
                "wednesday": "Wednesday",
                "thursday": "Thursday",
                "friday": "Friday",
            }
            found = []
            for key, name in day_map.items():
                pos = text.find(key)
                if pos != -1:
                    found.append((pos, name))
            found.sort(key=lambda x: x[0])
            return [name for _, name in found]
    return []


def parse_online_section(ws: Worksheet, merge_map: dict, label_row: int, header_row: int, start_row: int, end_row: int) -> list[dict]:
    """
    Online classes: single time header row, then odd/even sub-rows.
    No room column — room is always 'Online'.
    Day is detected from the label row (e.g. 'Online Classes (Saturday/ Wednesday)').
    """
    entries = []
    if start_row <= 0 or end_row <= 0 or header_row <= 0: return []
    
    time_slots = parse_online_time_headers(ws, header_row)
    online_days = _detect_online_days(ws, label_row)

    for row_idx in range(start_row, end_row + 1):
        # Prevent leakage of merged cell markers from upper sections
        week_marker_val = ws.cell(row_idx, 3).value
        if week_marker_val is None:
            for merge_range in ws.merged_cells.ranges:
                if row_idx in range(merge_range.min_row, merge_range.max_row + 1) and 3 in range(merge_range.min_col, merge_range.max_col + 1):
                    if merge_range.min_row >= start_row:
                        week_marker_val = ws.cell(merge_range.min_row, merge_range.min_col).value
                    break

        week_marker = str(week_marker_val or "").strip().lower()

        if ODD_MARKER in week_marker:
            week_note = "Odd weeks only"
            section_type = "online_odd"
        elif EVEN_MARKER in week_marker:
            week_note = "Even weeks only"
            section_type = "online_even"
        else:
            week_note = ""
            section_type = "online"

        base_props = {
            "room": ONLINE_ROOM,
            "section_type": section_type,
            "week_note": week_note,
        }

        def online_modifier(entry, col_idx):
            # Forcefully remove odd/even tag for CSE 1258 at the source
            if entry["course_code"].strip().upper().replace(" ", "") == "CSE1258":
                entry["section_type"] = "online"
                entry["week_note"] = ""

            # Look for a specific day in adjacent columns for this class
            specific_day = None
            day_names = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
            for offset in (1, 2, 3):
                adj_col = col_idx + offset
                # Look upwards from current row up to start_row
                for search_row in range(row_idx, start_row - 1, -1):
                    adj_val = str(ws.cell(search_row, adj_col).value or "").strip()
                    for d_name in day_names:
                        if d_name.lower() == adj_val.lower():
                            specific_day = d_name
                            break
                    if specific_day:
                        break
                if specific_day:
                    break
            
            # Determine online day for this column based on position
            col_online_day = None
            if len(online_days) >= 2:
                col_online_day = online_days[0] if col_idx < 11 else online_days[1]
            elif len(online_days) == 1:
                col_online_day = online_days[0]

            # Assign day: Specific row day > Header day
            if specific_day:
                entry["_online_day"] = specific_day
            elif col_online_day:
                entry["_online_day"] = col_online_day
                
            return entry

        row_entries = extract_row_entries(
            ws, 
            merge_map, 
            row_idx, 
            time_slots, 
            base_props, 
            col_start=1,
            entry_modifier=online_modifier
        )
        entries.extend(row_entries)

    return entries
